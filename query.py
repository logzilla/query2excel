#!/usr/bin/env python3
from dotenv import load_dotenv
import os
import argparse
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, numbers
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import time
from datetime import datetime

# Load environment variables
load_dotenv()

# Set up command line arguments
parser = argparse.ArgumentParser(description="Query LogZilla and generate Excel report.")
parser.add_argument("-v", "--verbose", help="Increase output verbosity.", action="store_true")
parser.add_argument("-d", "--debug", help="Show debug information.", action="store_true")
args = parser.parse_args()

# Environment variables for API key and URL
LOGZILLA_INSTANCE = os.getenv("LOGZILLA_INSTANCE")
API_KEY = os.getenv("API_KEY")

def debug_log(message):
    if args.debug:
        print(message)

def verbose_log(message):
    if args.verbose or args.debug:  # Debug implies verbosity
        print(message)

# Function to start the query
def start_query():
    # Load the query from query.json file
    with open('query.json', 'r') as file:
        data = json.load(file)  # Data is loaded from the file

    url = f"{LOGZILLA_INSTANCE}/api/query"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"token {API_KEY}"
    }
    debug_log(f"Request URL: {url}")
    debug_log(f"Request Headers: {json.dumps(headers, indent=4)}")
    debug_log(f"Request Body: {json.dumps(data, indent=4)}")

    response = requests.post(url, json=data, headers=headers)
    debug_log(f"Response Status Code: {response.status_code}")
    debug_log(f"Response Body: {response.text}")

    if response.status_code not in [200, 202]:
        verbose_log("Failed to start query due to an unexpected response.")
        exit(1)

    return response.json().get("query_id")


# Function to retrieve query results
def retrieve_results(query_id):
    max_attempts = 50
    attempt_delay = 10  # seconds to wait between attempts
    for attempt in range(1, max_attempts + 1):
        verbose_log(f"Attempt {attempt} to retrieve results for query ID {query_id}")
        url = f"{LOGZILLA_INSTANCE}/api/query/{query_id}"
        headers = {"Authorization": f"token {API_KEY}"}

        debug_log(f"Request URL: {url}")
        debug_log(f"Request Headers: {json.dumps(headers, indent=4)}")

        response = requests.get(url, headers=headers)
        debug_log(f"Response Status Code: {response.status_code}")
        debug_log(f"Response Body: {response.text}")

        response_json = response.json()
        if response_json.get('status') == "IN_PROGRESS":
            verbose_log("Query is still in progress. Waiting before next attempt...")
            time.sleep(attempt_delay)  # Wait before the next attempt
            continue  # Skip the rest of the current loop iteration
        elif 'results' in response_json:
            return response_json
        else:
            verbose_log("Failed to retrieve query results. Please try again later.")
            exit(1)

    verbose_log("Query results not ready after maximum attempts.")
    exit(1)
def create_excel_with_chart(data):
    # Convert the data into a DataFrame
    df = pd.json_normalize(data['results']['details'])

    # Convert Unix timestamps to human-readable date-time format
    df['Date'] = pd.to_datetime(df['ts_from'], unit='s').dt.strftime('%Y-%m-%d')

    # Keep the count as numeric for charting purposes
    df['Count'] = df['count']

    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['Date', 'Count'])

    # Write data rows
    for _, row in df.iterrows():
        ws.append([row['Date'], row['Count']])

    # Apply the number format to all cells in the 'Count' column except the header
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '[>=999950]0.00,,"M";[<=-999950]0.00,,"M";0.00,"K"'

    # Create a LineChart
    chart = LineChart()
    chart.title = "Event Count Over Time"
    chart.y_axis.title = 'Count'
    chart.x_axis.title = 'Date'

    # Data for the chart (using 'Count' for actual plotting)
    data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=False)  # Set titles_from_data to False

    # Categories (Dates) for the X-axis
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.set_categories(cats)

    # Remove the legend
    chart.legend = None  # This will remove the legend

    # Adding the chart to the worksheet
    ws.add_chart(chart, "E2")

    # Save the workbook
    excel_file_path = "report.xlsx"
    wb.save(excel_file_path)
if __name__ == "__main__":
    verbose_log("Starting query...")
    query_id = start_query()
    if query_id:
        verbose_log(f"Query ID: {query_id}")
        results = retrieve_results(query_id)
        if results:
            verbose_log("Creating Excel report with chart...")
            create_excel_with_chart(results)

